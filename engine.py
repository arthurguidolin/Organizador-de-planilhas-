"""
Organizador Inteligente de Planilhas - Engine
Motor de leitura, processamento e estilização profissional de planilhas.
Suporta arquivos .xlsx, .xls e .csv.
"""

import os
import re
import csv
import datetime
from pathlib import Path
from typing import Callable, Dict, Any, Optional, List, Tuple, TypedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


class ThemePalette(TypedDict):
    """Estrutura estrita de tipagem para as paletas de cores do organizador."""
    header_fill: str
    header_font: str
    zebra_fill: str
    border_color: str
    header_border: str
    accent: str


# Paletas de cores executivas / profissionais
THEMES: Dict[str, ThemePalette] = {
    "Azul Corporativo": {
        "header_fill": "1E3A8A",      # Deep Navy Blue
        "header_font": "FFFFFF",      # Branco
        "zebra_fill": "F1F5F9",       # Slate 100 suave
        "border_color": "CBD5E1",     # Slate 300
        "header_border": "0F172A",    # Slate 900
        "accent": "2563EB",
    },
    "Esmeralda Executivo": {
        "header_fill": "065F46",      # Emerald Green
        "header_font": "FFFFFF",
        "zebra_fill": "F0FDF4",       # Emerald 50
        "border_color": "BBF7D0",     # Emerald 200
        "header_border": "064E3B",
        "accent": "059669",
    },
    "Grafite Moderno": {
        "header_fill": "1F2937",      # Charcoal Gray
        "header_font": "FFFFFF",
        "zebra_fill": "F3F4F6",       # Gray 100
        "border_color": "D1D5DB",     # Gray 300
        "header_border": "111827",
        "accent": "4B5563",
    },
    "Roxo Tech": {
        "header_fill": "581C87",      # Deep Purple
        "header_font": "FFFFFF",
        "zebra_fill": "FAF5FF",       # Purple 50
        "border_color": "E9D5FF",     # Purple 200
        "header_border": "3B0764",
        "accent": "7C3AED",
    },
    "Oceano Petróleo": {
        "header_fill": "0E7490",      # Cyan / Teal escuro
        "header_font": "FFFFFF",
        "zebra_fill": "ECFEFF",       # Cyan 50
        "border_color": "BAE6FD",     # Sky 200
        "header_border": "155E75",
        "accent": "0891B2",
    },
    "Vinho Elegante": {
        "header_fill": "881337",      # Rose / Burgundy
        "header_font": "FFFFFF",
        "zebra_fill": "FFF1F2",       # Rose 50
        "border_color": "FECDD3",     # Rose 200
        "header_border": "4C0519",
        "accent": "BE123C",
    }
}

# Padrões para detecção inteligente de tipo por nome de coluna
CURRENCY_KEYWORDS = {
    "preco", "preço", "valor", "val", "total", "custo", "subtotal",
    "faturamento", "receita", "desconto", "lucro", "saldo", "pagamento",
    "salario", "salário", "taxa", "comissao", "comissão", "tarifa", "montante",
    "liquido", "líquido", "bruto", "vlr", "venda", "orcamento", "orçamento"
}

CODE_OR_ID_KEYWORDS = {
    "id", "codigo", "código", "cod", "cpf", "cnpj", "cep", "rg", "sku",
    "matricula", "matrícula", "telefone", "celular", "fone", "tel", "ano",
    "numero", "número", "num", "nota", "nf", "nfe", "pedido", "registro"
}

DATE_KEYWORDS = {
    "data", "dt", "date", "nascimento", "vencimento", "admissao", "admissão",
    "emissao", "emissão", "atualizacao", "atualização", "criacao", "criação"
}

PERCENT_KEYWORDS = {
    "porcentagem", "percentual", "%", "taxa_perc", "desconto_perc", "margem"
}


def normalize_text(text: str) -> str:
    """Normaliza texto removendo acentos e espaços para matching de keywords."""
    if not text:
        return ""
    text = str(text).lower().strip()
    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "õ": "o", "ô": "o",
        "ú": "u", "ü": "u",
        "ç": "c", "_": " ", "-": " "
    }
    for orig, rep in replacements.items():
        text = text.replace(orig, rep)
    return text


def detect_csv_delimiter_and_encoding(filepath: str) -> Tuple[str, str]:
    """Detecta automaticamente o delimitador e a codificação do arquivo CSV."""
    encodings_to_try = ["utf-8-sig", "utf-8", "latin1", "cp1252", "iso-8859-1"]
    sample_bytes = min(os.path.getsize(filepath), 65536)
    
    with open(filepath, "rb") as f:
        raw_data = f.read(sample_bytes)
        
    detected_encoding = "utf-8"
    for enc in encodings_to_try:
        try:
            text = raw_data.decode(enc)
            detected_encoding = enc
            break
        except (UnicodeDecodeError, LookupError):
            continue

    # Delimitadores comuns
    with open(filepath, "r", encoding=detected_encoding, errors="replace") as f:
        sample_text = f.read(sample_bytes)

    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=[",", ";", "\t", "|"])
        delimiter = dialect.delimiter
    except Exception:
        # Heurística caso o Sniffer falhe
        first_lines = sample_text.splitlines()[:5]
        delimiters = [";", ",", "\t", "|"]
        counts = {d: sum(line.count(d) for line in first_lines) for d in delimiters}
        delimiter = max(counts, key=counts.get)
        if counts[delimiter] == 0:
            delimiter = ","

    return delimiter, detected_encoding


class SpreadsheetOrganizer:
    """Motor de organização e estilização de planilhas."""

    # Expressões regulares pré-compiladas para alta performance
    RE_FLOAT_BR = re.compile(r"^-?\d{1,3}(\.\d{3})*,\d+$")
    RE_FLOAT_BR_SIMPLE = re.compile(r"^-?\d+,\d+$")
    RE_FLOAT_STD = re.compile(r"^-?\d+(\.\d+)?$")
    RE_DATE = re.compile(r"^(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})(?: \d{2}:\d{2}:\d{2})?$")
    RE_DATE_SHORT = re.compile(r"^(\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2})$")

    def __init__(self, options: Optional[Dict[str, Any]] = None):
        self.options = {
            "theme": "Azul Corporativo",
            "auto_column_width": True,
            "header_style": True,
            "freeze_panes": True,
            "auto_filters": True,
            "zebra_stripes": True,
            "smart_align": True,
            "smart_formats": True,
            "show_gridlines": True,
            "font_family": "Segoe UI",
            "output_suffix": "_organizado",
        }
        if options:
            self.options.update(options)

    def load_as_openpyxl_workbook(self, filepath: str) -> openpyxl.Workbook:
        """Carrega .xlsx, .xls ou .csv em um objeto openpyxl.Workbook seguro."""
        path = Path(filepath)
        ext = path.suffix.lower()

        if ext == ".xlsx":
            # Abre com openpyxl preservando dados
            return openpyxl.load_workbook(filepath, data_only=False)

        elif ext == ".xls":
            # Converte .xls via pandas (xlrd) para openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove aba padrão vazia
            
            excel_file = pd.ExcelFile(filepath, engine="xlrd")
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name)
                ws = wb.create_sheet(title=sheet_name[:31])
                
                # Escreve cabeçalho
                headers = list(df.columns)
                ws.append(headers)
                
                # Escreve linhas
                for row in df.itertuples(index=False):
                    clean_row = [None if pd.isna(val) else val for val in row]
                    ws.append(clean_row)
            return wb

        elif ext == ".csv":
            delimiter, encoding = detect_csv_delimiter_and_encoding(filepath)
            df = pd.read_csv(filepath, sep=delimiter, encoding=encoding, dtype=str)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            sheet_title = path.stem[:31]
            ws.title = sheet_title if sheet_title else "Dados"

            # Escreve cabeçalhos
            headers = list(df.columns)
            ws.append(headers)

            # Regexes pré-compiladas no escopo local para maximizar performance
            re_float_br = self.RE_FLOAT_BR
            re_float_br_simple = self.RE_FLOAT_BR_SIMPLE
            re_float_std = self.RE_FLOAT_STD
            re_date = self.RE_DATE

            # Tenta converter valores numéricos e datas para tipos adequados
            for row in df.itertuples(index=False):
                typed_row = []
                for val in row:
                    if pd.isna(val) or val == "" or str(val).strip() == "":
                        typed_row.append(None)
                    else:
                        val_str = str(val).strip()
                        # Tentativa de converter int/float
                        converted = False
                        # Checa se é float brasileiro (ex: 1.234,56 ou 1234,56)
                        if re_float_br.match(val_str):
                            try:
                                typed_row.append(float(val_str.replace(".", "").replace(",", ".")))
                                converted = True
                            except ValueError:
                                pass
                        elif re_float_br_simple.match(val_str):
                            try:
                                typed_row.append(float(val_str.replace(",", ".")))
                                converted = True
                            except ValueError:
                                pass
                        elif re_float_std.match(val_str):
                            try:
                                if "." in val_str:
                                    typed_row.append(float(val_str))
                                else:
                                    typed_row.append(int(val_str))
                                converted = True
                            except ValueError:
                                pass

                        if not converted:
                            # Checa se é data ISO ou brasileira com pré-validação rápida de tamanho e regex
                            val_len = len(val_str)
                            if val_len in (10, 19) and re_date.match(val_str):
                                date_parsed = False
                                formats = (
                                    ("%d/%m/%Y", "%Y-%m-%d")
                                    if val_len == 10
                                    else ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S")
                                )
                                for fmt in formats:
                                    try:
                                        dt = datetime.datetime.strptime(val_str, fmt)
                                        typed_row.append(dt)
                                        date_parsed = True
                                        break
                                    except ValueError:
                                        pass
                                if not date_parsed:
                                    typed_row.append(val_str)
                            else:
                                typed_row.append(val_str)
                ws.append(typed_row)
            return wb

        else:
            raise ValueError(f"Extensão não suportada: {ext}")

    def style_worksheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, options: Dict[str, Any]):
        """Aplica todas as regras de visual e organização em uma aba."""
        if ws.max_row == 0 or ws.max_column == 0:
            return

        theme_name = options.get("theme", "Azul Corporativo")
        theme = THEMES.get(theme_name, THEMES["Azul Corporativo"])
        font_family = options.get("font_family", "Segoe UI")

        header_font = Font(name=font_family, size=11, bold=True, color=theme["header_font"])
        header_fill = PatternFill(start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid")
        
        data_font = Font(name=font_family, size=10)
        zebra_fill = PatternFill(start_color=theme["zebra_fill"], end_color=theme["zebra_fill"], fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Bordas
        thin_border_side = Side(border_style="thin", color=theme["border_color"])
        cell_border = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side
        )

        header_border_side = Side(border_style="medium", color=theme["header_border"])
        header_border = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=header_border_side
        )

        # Mapeamento de colunas para formatação
        col_types = {}
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            raw_header = str(header_cell.value or "")
            norm_header = normalize_text(raw_header)
            
            # Detecta intenção pelo nome do cabeçalho
            col_type = "general"
            for kw in CURRENCY_KEYWORDS:
                if kw in norm_header:
                    col_type = "currency"
                    break
            if col_type == "general":
                for kw in PERCENT_KEYWORDS:
                    if kw in norm_header:
                        col_type = "percent"
                        break
            if col_type == "general":
                for kw in DATE_KEYWORDS:
                    if kw in norm_header:
                        col_type = "date"
                        break
            if col_type == "general":
                for kw in CODE_OR_ID_KEYWORDS:
                    if kw in norm_header or norm_header == kw:
                        col_type = "code"
                        break
            col_types[col_idx] = (col_type, raw_header)

        # 1. Estilização do Cabeçalho (Linha 1)
        if options.get("header_style", True):
            ws.row_dimensions[1].height = 28
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # 2. Estilização e Alinhamento das Linhas de Dados (Linha 2 em diante)
        max_row = ws.max_row
        max_col = ws.max_column

        for row_idx in range(2, max_row + 1):
            # Altura confortável da linha de dados
            ws.row_dimensions[row_idx].height = 21
            is_even = (row_idx % 2 == 0)

            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value

                # Fonte padrão
                cell.font = data_font
                cell.border = cell_border

                # Leitura zebrada
                if options.get("zebra_stripes", True) and is_even:
                    cell.fill = zebra_fill
                elif options.get("zebra_stripes", True):
                    cell.fill = white_fill

                # Classificação de alinhamento e formatação
                col_type, _ = col_types.get(col_idx, ("general", ""))
                
                if val is None:
                    continue

                # Formatações inteligentes e alinhamento
                if isinstance(val, (datetime.date, datetime.datetime)):
                    if options.get("smart_align", True):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    if options.get("smart_formats", True):
                        if isinstance(val, datetime.datetime) and (val.hour != 0 or val.minute != 0 or val.second != 0):
                            cell.number_format = "DD/MM/YYYY HH:MM:SS"
                        else:
                            cell.number_format = "DD/MM/YYYY"

                elif isinstance(val, bool):
                    if options.get("smart_align", True):
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                elif isinstance(val, (int, float)):
                    if col_type == "currency" and options.get("smart_formats", True):
                        cell.number_format = '"R$" #,##0.00'
                        if options.get("smart_align", True):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_type == "percent" and options.get("smart_formats", True):
                        cell.number_format = '0.00%'
                        if options.get("smart_align", True):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_type == "code":
                        cell.number_format = '0'
                        if options.get("smart_align", True):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        # Número padrão
                        if isinstance(val, float):
                            if options.get("smart_formats", True):
                                cell.number_format = '#,##0.00'
                            if options.get("smart_align", True):
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            if options.get("smart_formats", True):
                                cell.number_format = '#,##0'
                            if options.get("smart_align", True):
                                cell.alignment = Alignment(horizontal="right", vertical="center")

                elif isinstance(val, str):
                    val_clean = val.strip()
                    # Detecção de string numérica ou código
                    if col_type == "code" or (val_clean.isdigit() and len(val_clean) >= 6):
                        if options.get("smart_align", True):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif self.RE_DATE_SHORT.match(val_clean):
                        if options.get("smart_align", True):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        if options.get("smart_align", True):
                            cell.alignment = Alignment(horizontal="left", vertical="center")

        # 3. Ajuste Automático de Largura de Colunas
        if options.get("auto_column_width", True):
            max_check_row = min(max_row, 500)
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                
                for row_idx in range(1, max_check_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val is None:
                        continue

                    # Estimativa do tamanho renderizado
                    if isinstance(val, datetime.datetime):
                        length = 19
                    elif isinstance(val, datetime.date):
                        length = 10
                    elif isinstance(val, (int, float)):
                        col_type = col_types.get(col_idx, ("general", ""))[0]
                        if col_type == "currency":
                            length = len(f"R$ {val:,.2f}") + 3
                        else:
                            length = len(f"{val:,.2f}") + 1
                    else:
                        length = len(str(val))
                    
                    if length > max_len:
                        max_len = length

                    # Early break se já atingiu a largura máxima permitida (60)
                    if length >= 60:
                        max_len = 60
                        break

                # Margem de respiro + espaço para o botão do auto-filtro
                filter_padding = 4 if options.get("auto_filters", True) else 2
                calculated_width = max(max_len + filter_padding, 12)
                # Limite máximo razoável para evitar colunas gigantescas
                final_width = min(calculated_width, 60)
                ws.column_dimensions[col_letter].width = final_width

        # 4. Congelamento de Painel no Cabeçalho
        if options.get("freeze_panes", True):
            ws.freeze_panes = "A2"

        # 5. Filtros Automáticos
        if options.get("auto_filters", True) and max_row > 1 and max_col > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # 6. Exibição de Linhas de Grade
        if options.get("show_gridlines", True) and ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True

    def process_file(
        self,
        input_path: str,
        output_dir: Optional[str] = None,
        progress_callback: Optional[Callable[[str, float], None]] = None
    ) -> str:
        """
        Processa um arquivo, aplica a estilização e salva um novo arquivo _organizado.xlsx.
        Garante que o arquivo de origem seja aberto em modo somente leitura e permaneça 100% intacto.
        """
        in_path = Path(input_path).resolve()
        if not in_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {input_path}")

        if output_dir:
            out_dir = Path(output_dir).resolve()
            out_dir.mkdir(parents=True, exist_ok=True)
        else:
            out_dir = in_path.parent

        suffix = self.options.get("output_suffix", "_organizado")
        target_filename = f"{in_path.stem}{suffix}.xlsx"
        out_path = out_dir / target_filename

        if progress_callback:
            progress_callback(f"Carregando '{in_path.name}' com segurança...", 0.2)

        # 1. Carrega dados de forma não destrutiva
        wb = self.load_as_openpyxl_workbook(str(in_path))

        # 2. Processa todas as abas
        sheet_names = wb.sheetnames
        total_sheets = len(sheet_names)

        for idx, sheet_name in enumerate(sheet_names, start=1):
            if progress_callback:
                pct = 0.2 + (0.6 * (idx / total_sheets))
                progress_callback(f"Estilizando aba '{sheet_name}' ({idx}/{total_sheets})...", pct)
            ws = wb[sheet_name]
            self.style_worksheet(ws, self.options)

        # 3. Salva novo arquivo
        if progress_callback:
            progress_callback(f"Salvando planilha organizada em '{out_path.name}'...", 0.9)

        wb.save(str(out_path))

        if progress_callback:
            progress_callback(f"Concluído com sucesso: '{out_path.name}'", 1.0)

        return str(out_path)
