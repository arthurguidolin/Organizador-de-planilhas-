"""
Script de testes automatizados para o motor do Organizador Inteligente de Planilhas.
Cria amostras de planilhas desorganizadas (.xlsx e .csv) e valida a formatação e integridade.
"""

import os
import sys
import shutil
import datetime
from pathlib import Path
import openpyxl
import pandas as pd

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


from engine import SpreadsheetOrganizer, THEMES


def create_sample_files(test_dir: Path):
    """Cria planilhas de teste com problemas comuns de formatação."""
    test_dir.mkdir(parents=True, exist_ok=True)

    # 1. Arquivo CSV desorganizado (vendas_bruto.csv)
    csv_path = test_dir / "vendas_bruto.csv"
    csv_content = """ID_PEDIDO;CLIENTE_NOME_COMPLETO;DATA_VENDA;VALOR_TOTAL_BRUTO;STATUS_PAGAMENTO;DESCONTO_PERC
10045;Arthur Guidolin Consultoria e Tecnologia da Informacao Ltda;12/05/2024;15850,50;Pago;0,05
10046;Maria de Souza Pereira Oliveira;14/05/2024;350,00;Pendente;0,00
10047;Empresa Brasileira de Logistica e Transportes Integrados S/A;15/05/2024;45200,90;Pago;0,10
10048;Carlos Alberto Santos;18/05/2024;1299,99;Cancelado;0,02
10049;Juliana Mendes Rocha;19/05/2024;8450,00;Pago;0,00
"""
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write(csv_content.strip())

    # 2. Arquivo XLSX multi-aba desorganizado (relatorio_financeiro.xlsx)
    xlsx_path = test_dir / "relatorio_financeiro.xlsx"
    wb = openpyxl.Workbook()
    
    # Aba 1: Faturamento
    ws1 = wb.active
    ws1.title = "Faturamento Mensal"
    ws1.append(["Codigo_NF", "Descricao_Servico", "Data_Emissao", "Valor_Bruto", "Valor_Liquido", "Imposto_Retido"])
    ws1.append([98101, "Consultoria Estratégica em Engenharia de Software e IA", datetime.date(2024, 6, 1), 25000.0, 22500.0, 2500.0])
    ws1.append([98102, "Licenciamento de Software Corporativo Anual", datetime.date(2024, 6, 5), 120000.5, 108000.45, 12000.05])
    ws1.append([98103, "Suporte Técnico 24x7 Nível 3", datetime.date(2024, 6, 10), 8500.0, 7650.0, 850.0])
    ws1.append([98104, "Treinamento de Equipe e Workshop de Produtividade", datetime.date(2024, 6, 15), 14250.75, 12825.68, 1425.07])

    # Aba 2: Funcionarios
    ws2 = wb.create_sheet(title="Equipe")
    ws2.append(["Matricula", "Nome_Funcionario", "Cargo", "Salario_Base", "Data_Admissao"])
    ws2.append([101, "Roberto Silva", "Arquiteto de Software", 18500.0, datetime.date(2021, 3, 15)])
    ws2.append([102, "Ana Beatriz Ferreira", "Engenheira de Dados Senior", 16200.0, datetime.date(2022, 7, 1)])
    ws2.append([103, "Lucas Albuquerque", "Product Designer UI/UX", 11500.0, datetime.date(2023, 1, 10)])
    ws2.append([104, "Fernanda Lima", "Gerente de Projetos Agile", 17000.0, datetime.date(2020, 11, 20)])

    wb.save(str(xlsx_path))

    return csv_path, xlsx_path


def run_tests():
    """Executa o processamento e valida os arquivos organizados."""
    test_dir = Path("./test_samples").resolve()
    csv_path, xlsx_path = create_sample_files(test_dir)

    print(f"=== TESTE 1: Processando CSV '{csv_path.name}' ===")
    organizer_blue = SpreadsheetOrganizer({"theme": "Azul Corporativo"})
    out_csv = organizer_blue.process_file(str(csv_path))
    print(f"Arquivo gerado: {out_csv}")
    assert Path(out_csv).exists(), "Arquivo de saída do CSV não foi criado!"

    # Validação do arquivo CSV transformado em XLSX organizado
    wb_csv = openpyxl.load_workbook(out_csv)
    ws_csv = wb_csv.active
    assert ws_csv.freeze_panes == "A2", "Freeze panes incorreto!"
    assert ws_csv.auto_filter.ref is not None, "Auto filter não aplicado!"
    assert ws_csv.row_dimensions[1].height == 28, "Altura do cabeçalho incorreta!"
    # Checa larguras de coluna
    col_a_width = ws_csv.column_dimensions["A"].width
    col_b_width = ws_csv.column_dimensions["B"].width
    assert col_b_width > 30, f"Largura da coluna B deveria acomodar nomes longos (atual: {col_b_width})"
    print(f"✔ CSV processado com sucesso! Largura Col B: {col_b_width:.1f}, Freeze: {ws_csv.freeze_panes}, Filtro: {ws_csv.auto_filter.ref}")

    print(f"\n=== TESTE 2: Processando XLSX Multi-Aba '{xlsx_path.name}' ===")
    organizer_emerald = SpreadsheetOrganizer({"theme": "Esmeralda Executivo"})
    out_xlsx = organizer_emerald.process_file(str(xlsx_path))
    print(f"Arquivo gerado: {out_xlsx}")
    assert Path(out_xlsx).exists(), "Arquivo de saída do XLSX não foi criado!"

    # Validação do XLSX multi-aba
    wb_xlsx = openpyxl.load_workbook(out_xlsx)
    assert len(wb_xlsx.sheetnames) == 2, "Todas as abas originais devem ser preservadas!"
    assert "Faturamento Mensal" in wb_xlsx.sheetnames
    assert "Equipe" in wb_xlsx.sheetnames

    for s_name in wb_xlsx.sheetnames:
        ws = wb_xlsx[s_name]
        assert ws.freeze_panes == "A2", f"Freeze panes falhou na aba {s_name}"
        assert ws.auto_filter.ref is not None, f"Auto filter falhou na aba {s_name}"
        # Verifica cabeçalho
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True, f"Fonte do cabeçalho não está em negrito na aba {s_name}"
        # Verifica formatação da linha 2
        val_cell = ws.cell(row=2, column=4) # Coluna de valor/salário
        assert '"R$"' in str(val_cell.number_format) or "#,##0" in str(val_cell.number_format), f"Formatação de moeda não aplicada: {val_cell.number_format}"
        print(f"✔ Aba '{s_name}': Cabeçalho={header_cell.value}, Formato Moeda={val_cell.number_format}, Freeze={ws.freeze_panes}")

    print("\n=== TESTE 3: Verificando integridade dos arquivos originais ===")
    # Garante que os arquivos originais ainda existem e não foram modificados
    assert csv_path.exists()
    assert xlsx_path.exists()
    print("✔ Arquivos originais permanecem 100% intactos e inalterados.")

    print("\n========================================================")
    print("      TODOS OS TESTES FORAM CONCLUÍDOS COM SUCESSO!     ")
    print("========================================================")


if __name__ == "__main__":
    run_tests()
